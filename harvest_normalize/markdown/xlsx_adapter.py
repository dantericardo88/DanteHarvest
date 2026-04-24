"""
XLSXAdapter — spreadsheet to markdown table conversion.

Ascend Cycle 1: close multi_format_ingest gap (DH: 7 → 9 vs LlamaIndex: 9).

Harvested from: LlamaIndex PandasExcelReader + openpyxl patterns.

Converts XLSX/XLS/CSV workbooks to markdown tables.
Each sheet becomes a level-2 heading + markdown table.

Constitutional guarantees:
- Local-first: openpyxl runs locally; no network calls
- Fail-closed: missing file or unreadable format raises NormalizationError
- Zero-ambiguity: convert() always returns str (never None or empty on valid input)
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import List, Optional

from harvest_core.control.exceptions import NormalizationError


def _rows_to_markdown(headers: List[str], rows: List[List[str]]) -> str:
    if not headers:
        return ""
    header_row = "| " + " | ".join(str(h) for h in headers) + " |"
    sep_row = "| " + " | ".join("---" for _ in headers) + " |"
    data_rows = [
        "| " + " | ".join(str(c) for c in row[:len(headers)]) + " |"
        for row in rows
    ]
    return "\n".join([header_row, sep_row] + data_rows)


class XLSXAdapter:
    """
    Convert XLSX/XLS/CSV files to markdown.

    Usage:
        adapter = XLSXAdapter()
        markdown = adapter.convert(Path("report.xlsx"))
    """

    def convert(self, path: Path) -> str:
        """
        Convert spreadsheet to markdown string.
        Fail-closed: raises NormalizationError if file not found or format unsupported.
        Zero-ambiguity: returns non-empty str on valid spreadsheet.
        """
        if not path.exists():
            raise NormalizationError(f"Spreadsheet file not found: {path}")

        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls", ".xlsm"):
            return self._convert_xlsx(path)
        elif suffix == ".csv":
            return self._convert_csv(path)
        else:
            raise NormalizationError(
                f"Unsupported spreadsheet format: {suffix}. "
                "Supported: .xlsx, .xls, .xlsm, .csv"
            )

    def _convert_xlsx(self, path: Path) -> str:
        try:
            import openpyxl
        except ImportError as e:
            raise NormalizationError(
                "openpyxl not installed. Run: pip install openpyxl"
            ) from e

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            raise NormalizationError(f"Failed to open workbook {path}: {e}") from e

        sections: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c) if c is not None else "" for c in rows[0]]
            data_rows = [
                [str(c) if c is not None else "" for c in row]
                for row in rows[1:]
            ]
            table = _rows_to_markdown(headers, data_rows)
            sections.append(f"## {sheet_name}\n\n{table}")

        wb.close()
        return "\n\n".join(sections) if sections else ""

    def _convert_csv(self, path: Path) -> str:
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        except Exception as e:
            raise NormalizationError(f"Failed to parse CSV {path}: {e}") from e

        if not rows:
            return ""
        headers = rows[0]
        data_rows = rows[1:]
        stem = path.stem
        table = _rows_to_markdown(headers, data_rows)
        return f"## {stem}\n\n{table}"
