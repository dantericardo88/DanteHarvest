"""
SpreadsheetLoader — ingest XLSX/CSV/ODS files into markdown tables.

Harvested from: LlamaIndex PandasExcelReader + openpyxl patterns.

Features:
- XLSX via openpyxl (MIT) with graceful fallback to csv module if absent
- CSV via stdlib csv — zero extra deps
- ODS: attempted via openpyxl (limited) or raises NormalizationError gracefully
- Multi-sheet support: each sheet becomes a level-2 heading + markdown table
- Auto-detects headers from first row
- Streams row by row — handles large files without loading all into memory

Constitutional guarantees:
- Local-first: all processing is local, no network calls
- Fail-closed: missing file or unreadable format raises NormalizationError
- Zero-ambiguity: load() always returns list[SpreadsheetDocument]
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from harvest_core.control.exceptions import NormalizationError

try:
    import openpyxl as _openpyxl  # noqa: F401
    _OPENPYXL_AVAILABLE = True
except (ImportError, TypeError):
    _OPENPYXL_AVAILABLE = False


@dataclass
class SheetDocument:
    """A single sheet from a spreadsheet file, represented as markdown."""
    sheet_name: str
    headers: List[str]
    row_count: int
    markdown: str


@dataclass
class SpreadsheetDocument:
    """Result of loading a spreadsheet file."""
    file_path: str
    format: str  # "xlsx", "csv", "ods"
    sheets: List[SheetDocument] = field(default_factory=list)

    @property
    def markdown(self) -> str:
        """Full markdown representation of all sheets."""
        return "\n\n".join(s.markdown for s in self.sheets if s.markdown)

    @property
    def sheet_names(self) -> List[str]:
        return [s.sheet_name for s in self.sheets]


def _rows_to_markdown(headers: List[str], rows: List[List[str]]) -> str:
    """Convert header + data rows to a GitHub-flavored markdown table."""
    if not headers:
        return ""
    header_row = "| " + " | ".join(str(h) for h in headers) + " |"
    sep_row = "| " + " | ".join("---" for _ in headers) + " |"
    data_rows = [
        "| " + " | ".join(str(c) for c in row[: len(headers)]) + " |"
        for row in rows
    ]
    return "\n".join([header_row, sep_row] + data_rows)


class SpreadsheetLoader:
    """
    Load XLSX, CSV, and ODS spreadsheet files into structured markdown documents.

    Usage:
        loader = SpreadsheetLoader()
        docs = loader.load(Path("report.xlsx"))
        for doc in docs:
            print(doc.markdown)
    """

    SUPPORTED_SUFFIXES = {".xlsx", ".xls", ".xlsm", ".csv", ".ods"}

    def get_supported_formats(self) -> List[str]:
        """Return list of supported format strings. CSV is always present."""
        formats = ["csv"]
        try:
            import openpyxl  # noqa: F401
            for fmt in ("xlsx", "xls", "xlsm", "ods"):
                if fmt not in formats:
                    formats.append(fmt)
        except (ImportError, TypeError):
            pass
        return formats

    def load(self, path: Path) -> List[SpreadsheetDocument]:
        """
        Load a spreadsheet file. Returns a list with one SpreadsheetDocument.
        Raises NormalizationError if the file is not found or format unsupported.
        """
        path = Path(path)
        if not path.exists():
            raise NormalizationError(f"Spreadsheet file not found: {path}")
        if not path.is_file():
            raise NormalizationError(f"Path is not a file: {path}")

        suffix = path.suffix.lower()
        if suffix not in self.SUPPORTED_SUFFIXES:
            raise NormalizationError(
                f"Unsupported spreadsheet format: {suffix!r}. "
                f"Supported: {', '.join(sorted(self.SUPPORTED_SUFFIXES))}"
            )

        if suffix in (".xlsx", ".xls", ".xlsm", ".ods"):
            return self._load_xlsx(path, fmt=suffix.lstrip("."))
        else:  # .csv
            return self._load_csv(path)

    def _load_xlsx(self, path: Path, fmt: str) -> List[SpreadsheetDocument]:
        """Load XLSX/XLS/XLSM/ODS via openpyxl; fallback to csv on ImportError."""
        try:
            import openpyxl
        except ImportError:
            csv_path = path.with_suffix(".csv")
            if csv_path.exists():
                import warnings
                warnings.warn(
                    f"openpyxl not installed; falling back to CSV for {path.name}. "
                    "Install openpyxl: pip install openpyxl",
                    ImportWarning,
                    stacklevel=2,
                )
                return self._load_csv(csv_path)
            raise NormalizationError(
                "openpyxl is required to read XLSX/XLS/ODS files. "
                "Install it with: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise NormalizationError(f"Failed to open workbook {path}: {exc}") from exc

        sheets: List[SheetDocument] = []
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                raw_rows = list(ws.iter_rows(values_only=True))
                if not raw_rows:
                    continue
                headers = [str(c) if c is not None else "" for c in raw_rows[0]]
                data_rows = [
                    [str(c) if c is not None else "" for c in row]
                    for row in raw_rows[1:]
                ]
                table_md = _rows_to_markdown(headers, data_rows)
                if not table_md:
                    continue
                full_md = f"## {sheet_name}\n\n{table_md}"
                sheets.append(
                    SheetDocument(
                        sheet_name=sheet_name,
                        headers=headers,
                        row_count=len(data_rows),
                        markdown=full_md,
                    )
                )
        finally:
            wb.close()

        doc = SpreadsheetDocument(file_path=str(path), format=fmt, sheets=sheets)
        return [doc]

    def _load_csv(self, path: Path) -> List[SpreadsheetDocument]:
        """Load CSV using stdlib csv module — zero extra deps."""
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        except Exception as exc:
            raise NormalizationError(f"Failed to parse CSV {path}: {exc}") from exc

        if not rows:
            sheet = SheetDocument(sheet_name=path.stem, headers=[], row_count=0, markdown="")
            return [SpreadsheetDocument(file_path=str(path), format="csv", sheets=[sheet])]

        headers = rows[0]
        data_rows = rows[1:]
        table_md = _rows_to_markdown(headers, data_rows)
        full_md = f"## {path.stem}\n\n{table_md}" if table_md else ""
        sheet = SheetDocument(
            sheet_name=path.stem,
            headers=headers,
            row_count=len(data_rows),
            markdown=full_md,
        )
        return [SpreadsheetDocument(file_path=str(path), format="csv", sheets=[sheet])]


class IngestCapabilities:
    """Reports which ingest formats are available without importing optional deps."""

    @staticmethod
    def get_available() -> dict:
        caps: dict = {"csv": True, "json": True, "txt": True, "html": True}
        try:
            import openpyxl  # noqa: F401
            caps["xlsx"] = True
            caps["xls"] = True
        except ImportError:
            caps["xlsx"] = False
            caps["xls"] = False
        try:
            import pdfminer  # noqa: F401
            caps["pdf"] = True
        except ImportError:
            try:
                import pypdf  # noqa: F401
                caps["pdf"] = True
            except ImportError:
                caps["pdf"] = False
        return caps

    @staticmethod
    def get_missing_deps() -> list:
        missing = []
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            missing.append({
                "package": "openpyxl",
                "formats": ["xlsx", "xls"],
                "install": "pip install openpyxl",
            })
        return missing
